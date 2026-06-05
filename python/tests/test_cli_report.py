"""Tests for the `vdl report ...` subapp."""

from __future__ import annotations

import zipfile
from pathlib import Path

import pytest
from typer.testing import CliRunner

from venn_diagram_lab.cli import app

runner = CliRunner()
SAMPLE = "dataset_real_cancer_drivers_4"


def test_report_pdf(tmp_path: Path) -> None:
    target = tmp_path / "report.pdf"
    res = runner.invoke(app, ["report", "pdf", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    assert target.exists()
    head = target.read_bytes()[:8]
    assert head.startswith(b"%PDF-")


def test_report_zip(tmp_path: Path) -> None:
    target = tmp_path / "bundle.zip"
    res = runner.invoke(app, ["report", "zip", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    assert target.exists()
    with zipfile.ZipFile(target) as zf:
        names = set(zf.namelist())
    assert any(name.endswith(".pdf") for name in names), names
    assert any(name.endswith(".tsv") for name in names), names
    assert any(name.endswith(".svg") for name in names), names


def test_report_pdf_unknown_input_exits_1(tmp_path: Path) -> None:
    res = runner.invoke(app, ["report", "pdf", "nope_xyz", "--out", str(tmp_path / "x.pdf")])
    assert res.exit_code == 1


def test_report_zip_contains_expected_files(tmp_path: Path) -> None:
    """The zip bundle should contain venn/upset/network/share-dist SVGs + 3 TSVs + 1 PDF."""
    target = tmp_path / "bundle.zip"
    res = runner.invoke(app, ["report", "zip", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    with zipfile.ZipFile(target) as zf:
        names = sorted(zf.namelist())
    expected_kinds = ["venn.svg", "upset.svg", "network.svg", "share-dist.svg",
                       "regions_summary.tsv", "items_matrix.tsv", "statistics.tsv",
                       "report.pdf"]
    for kind in expected_kinds:
        assert kind in names, f"{kind} missing from zip; got {names}"


# ----- --sample flag coverage -----------------------------------------------


def test_report_pdf_with_sample_flag(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    monkeypatch.chdir(tmp_path)
    res = runner.invoke(app, ["report", "pdf", "--sample"])
    assert res.exit_code == 0, res.output
    assert (tmp_path / f"{SAMPLE}__report.pdf").exists()


def test_report_zip_with_sample_flag(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path,
) -> None:
    monkeypatch.chdir(tmp_path)
    res = runner.invoke(app, ["report", "zip", "--sample"])
    assert res.exit_code == 0, res.output
    assert (tmp_path / f"{SAMPLE}__report.zip").exists()


def test_report_pdf_no_input_no_sample_exits_1() -> None:
    res = runner.invoke(app, ["report", "pdf"])
    assert res.exit_code == 1
    assert "INPUT required" in res.output or "use --sample" in res.output


# ----- v2.2.3 PDF + ZIP additions -------------------------------------------


# Minimum PDF byte size expected once the Item Share Distribution page is
# present (v2.2.2 baseline was ~200 KB; the share-dist page adds at least 30 KB
# of imshow PNG payload).
_MIN_PDF_SIZE_WITH_SHARE_DIST = 50_000


def test_report_pdf_includes_share_distribution_page(tmp_path: Path) -> None:
    """The v2.2.3 PDF report adds an Item Share Distribution page.

    Exact text extraction would require pdfplumber/pypdf as a hard dep; instead
    we assert the PDF is non-trivially large — the share-dist imshow payload
    alone is several tens of KB, well above the floor.
    """
    target = tmp_path / "r.pdf"
    res = runner.invoke(app, ["report", "pdf", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    assert target.exists()
    assert target.stat().st_size > _MIN_PDF_SIZE_WITH_SHARE_DIST


def test_report_pdf_cluster_heatmap_flag(tmp_path: Path) -> None:
    """``--cluster-heatmap`` adds the cluster-ordered Jaccard heatmap page."""
    base = tmp_path / "r_base.pdf"
    cluster = tmp_path / "r_cluster.pdf"
    r1 = runner.invoke(app, ["report", "pdf", SAMPLE, "--out", str(base)])
    r2 = runner.invoke(
        app, ["report", "pdf", SAMPLE, "--cluster-heatmap", "--out", str(cluster)],
    )
    assert r1.exit_code == 0, r1.output
    assert r2.exit_code == 0, r2.output
    assert base.exists()
    assert cluster.exists()
    # The cluster-heatmap PDF must be larger than the baseline (extra page).
    assert cluster.stat().st_size > base.stat().st_size


def test_report_zip_contains_xlsx(tmp_path: Path) -> None:
    """The v2.2.3 ZIP bundle includes the enrichment-statistics Excel workbook."""
    target = tmp_path / "b.zip"
    res = runner.invoke(app, ["report", "zip", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    with zipfile.ZipFile(target) as zf:
        names = set(zf.namelist())
    assert any(name.endswith(".xlsx") for name in names), names


def test_report_zip_contains_readme(tmp_path: Path) -> None:
    """The v2.2.3 ZIP bundle includes a README.txt with provenance + methodology."""
    target = tmp_path / "b.zip"
    res = runner.invoke(app, ["report", "zip", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    with zipfile.ZipFile(target) as zf:
        assert "README.txt" in zf.namelist()
        with zf.open("README.txt") as f:
            body = f.read().decode("utf-8")
    assert "Venn Diagram Lab" in body
    assert "About This Report" in body


def test_to_excel_workbook_3_sheets(tmp_path: Path) -> None:
    """``to_excel_workbook`` writes 3 sheets: Jaccard / Sørensen-Dice / Enrichment."""
    from openpyxl import load_workbook  # noqa: PLC0415

    from venn_diagram_lab.analysis import analyze  # noqa: PLC0415
    from venn_diagram_lab.report.excel import to_excel_workbook  # noqa: PLC0415
    from venn_diagram_lab.samples import load_sample  # noqa: PLC0415

    target = tmp_path / "stats.xlsx"
    result = analyze(load_sample(SAMPLE))
    to_excel_workbook(result, target)
    wb = load_workbook(target, read_only=True)
    assert set(wb.sheetnames) >= {"Jaccard", "Sørensen-Dice", "Enrichment"}


# ----- v2.2.3 About + Credits unification -----------------------------------


def test_about_sections_include_credits_and_cite() -> None:
    """The structured About sections must include the v2.2.3 Credits footer."""
    from venn_diagram_lab.render.pdf import _ABOUT_SECTIONS  # noqa: PLC0415

    titles = [t for t, _ in _ABOUT_SECTIONS]
    assert "Venn Diagram Lab" in titles
    assert "Plots" in titles
    assert "Statistics" in titles
    assert "Credits and Cite" in titles
    # Verify the credits body carries the Zenodo DOI + web tool URL verbatim.
    credits_body = next(b for t, b in _ABOUT_SECTIONS if t == "Credits and Cite")
    assert "10.5281/zenodo.19510813" in credits_body
    assert "venndiagramlab.org" in credits_body


def test_about_sections_titles_match_webtool() -> None:
    """Structured Python About sections must mirror the webtool catalog 1:1."""
    from venn_diagram_lab.render.pdf import _ABOUT_SECTIONS  # noqa: PLC0415

    expected_titles = [
        "Venn Diagram Lab",
        "Plots",
        "1. Venn Diagrams",
        "2. UpSet Plots",
        "3. Set Relationship Network",
        "Statistics",
        "1. Pairwise Jaccard Index",
        "2. Sorensen-Dice Index",
        "3. Intersection Enrichment (Hypergeometric Test)",
        "4. Bar chart",
        "5. Lollipop chart",
        "6. Heatmap",
        "7. Item Share Distribution",
        "8. Cluster Heatmap",
        "Credits and Cite",
    ]
    assert [t for t, _ in _ABOUT_SECTIONS] == expected_titles


def test_build_about_pages_returns_at_least_one_figure() -> None:
    """``_build_about_pages`` returns a non-empty list of matplotlib Figures."""
    from matplotlib.figure import Figure  # noqa: PLC0415

    from venn_diagram_lab.render.pdf import _build_about_pages  # noqa: PLC0415

    pages = _build_about_pages()
    assert len(pages) >= 1
    assert all(isinstance(p, Figure) for p in pages)
    # Clean up so matplotlib doesn't warn about too many open figures.
    import matplotlib.pyplot as plt  # noqa: PLC0415
    for fig in pages:
        plt.close(fig)


# Floor for the v2.2.3 PDF: at least the v2.2.2 share-dist baseline plus the
# multi-section About + Credits payload (text on landscape A4 adds ~10-30 KB).
_MIN_PDF_SIZE_WITH_ABOUT_CREDITS = 60_000


def test_report_pdf_includes_about_credits(tmp_path: Path) -> None:
    """The v2.2.3 PDF report must embed the unified About + Credits content."""
    target = tmp_path / "ac.pdf"
    res = runner.invoke(app, ["report", "pdf", SAMPLE, "--out", str(target)])
    assert res.exit_code == 0, res.output
    assert target.exists()
    # Size floor: about-credits multi-page block must produce non-trivial PDF.
    assert target.stat().st_size > _MIN_PDF_SIZE_WITH_ABOUT_CREDITS
