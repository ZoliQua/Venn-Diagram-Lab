"""3-sheet Excel workbook matching the webtool's ZIP-bundle statistics file.

The webtool ships an ``enrichment_statistics_{n}-sets.xlsx`` workbook
alongside its *Download Everything* ZIP bundle. This module reproduces that
workbook from a :class:`venn_diagram_lab.analysis.RegionResult` using
``openpyxl`` directly (no xlsxwriter / pandas-Excel detour).
"""

from __future__ import annotations

from pathlib import Path
from typing import TYPE_CHECKING, Any, cast

if TYPE_CHECKING:
    import pandas as pd

    from venn_diagram_lab.analysis import RegionResult

__all__ = ["to_excel_workbook"]

# Sheet titles — match the webtool's three-sheet layout.
_SHEET_JACCARD = "Jaccard"
_SHEET_DICE = "Sørensen-Dice"
_SHEET_ENRICHMENT = "Enrichment"

# Header styling (mirrors the webtool's bold + grey header band).
_HEADER_FILL = "DDDDDD"


def _write_square_matrix(
    ws: Any,  # openpyxl Worksheet (no stubs in deps)
    set_names: list[str],
    matrix: pd.DataFrame,
) -> None:
    """Write an NxN pairwise metric matrix to ``ws``.

    Layout (1-indexed openpyxl coordinates):
      A1            = ""
      B1..(N+1)1    = set names (column headers)
      A2..A(N+1)    = set names (row headers)
      B2..(N+1)(N+1)= matrix values, 4 decimals.

    Header cells (row 1 + column A) are bold with a grey fill.
    """
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)

    # Column headers.
    for col_idx, name in enumerate(set_names, start=2):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill

    # Row headers + body.
    for row_idx, row_name in enumerate(set_names, start=2):
        hdr = ws.cell(row=row_idx, column=1, value=row_name)
        hdr.font = header_font
        hdr.fill = header_fill
        for col_idx, col_name in enumerate(set_names, start=2):
            raw = cast(float, matrix.loc[row_name, col_name])
            value = float(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.number_format = "0.0000"


def _write_enrichment_sheet(
    ws: Any,  # openpyxl Worksheet (no stubs in deps)
    result: RegionResult,
) -> None:
    """Write the long-form Enrichment sheet (one row per pair).

    Columns: ``set_a``, ``set_b``, ``intersection``, ``union``,
    ``expected``, ``fold_enrichment``, ``p_value``, ``fdr``,
    ``significant``.
    """
    from openpyxl.styles import Font, PatternFill  # noqa: PLC0415

    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL)

    headers = [
        "set_a", "set_b", "intersection", "union",
        "expected", "fold_enrichment", "p_value", "fdr", "significant",
    ]
    for col_idx, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill

    set_sizes = result.set_sizes
    stats = result.statistics

    for row_idx, (_, row) in enumerate(stats.hypergeometric.iterrows(), start=2):
        a = str(row["set_a"])
        b = str(row["set_b"])
        inter = int(row["intersection"])
        union = set_sizes[a] + set_sizes[b] - inter
        # fold_enrichment is in the square DataFrame.
        fe = float(cast(float, stats.fold_enrichment.loc[a, b]))
        ws.cell(row=row_idx, column=1, value=a)
        ws.cell(row=row_idx, column=2, value=b)
        ws.cell(row=row_idx, column=3, value=inter)
        ws.cell(row=row_idx, column=4, value=union)
        ws.cell(row=row_idx, column=5, value=float(row["expected"])).number_format = "0.00"
        ws.cell(row=row_idx, column=6, value=fe).number_format = "0.0000"
        ws.cell(row=row_idx, column=7, value=float(row["p_value"]))
        ws.cell(row=row_idx, column=8, value=float(row["p_adjusted"]))
        ws.cell(row=row_idx, column=9, value=bool(row["significant"]))


def to_excel_workbook(result: RegionResult, path: Path | str) -> None:
    """Write a 3-sheet Excel workbook matching the webtool's ZIP bundle.

    Sheets:
      - ``Jaccard`` — square matrix (sets x sets) of Jaccard indices.
      - ``Sørensen-Dice`` — square matrix of Sørensen-Dice coefficients.
      - ``Enrichment`` — long-form: ``set_a``, ``set_b``, ``intersection``,
        ``union``, ``expected``, ``fold_enrichment``, ``p_value``, ``fdr``,
        ``significant``.

    Uses ``openpyxl`` directly (pulled in as a pandas transitive dep, made
    explicit in :file:`pyproject.toml`).
    """
    from openpyxl import Workbook  # noqa: PLC0415

    set_names = list(result.dataset.set_names)
    stats = result.statistics

    wb = Workbook()
    # First sheet — rename the default sheet rather than removing+adding.
    ws_jacc = wb.active
    ws_jacc.title = _SHEET_JACCARD
    _write_square_matrix(ws_jacc, set_names, stats.jaccard)

    ws_dice = wb.create_sheet(_SHEET_DICE)
    _write_square_matrix(ws_dice, set_names, stats.dice)

    ws_enr = wb.create_sheet(_SHEET_ENRICHMENT)
    _write_enrichment_sheet(ws_enr, result)

    wb.save(str(Path(path)))
